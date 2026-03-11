import re
import time
import random
from typing import Optional, Dict, Tuple, List

import openpyxl
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError


URL = "https://www.idealista.com/sala-de-prensa/informes-precio-vivienda/venta/andalucia/malaga-provincia/"

MUNICIPIOS = [
    "alhaurin_de_la_torre",
    "benahavis",
    "benalmadena",
    "estepona",
    "fuengirola",
    "marbella",
    "mijas",
    "rincon_de_la_victoria",
    "torremolinos",
    "malaga",
]

ANHOS = list(range(2015, 2025))  # 2015-2024
TRIMESTRES = ["Q1", "Q2", "Q3", "Q4"]

MUNICIPIO_LABEL = {
    "alhaurin_de_la_torre": "Alhaurín de la Torre",
    "benahavis": "Benahavís",
    "benalmadena": "Benalmádena",
    "estepona": "Estepona",
    "fuengirola": "Fuengirola",
    "marbella": "Marbella",
    "mijas": "Mijas",
    "rincon_de_la_victoria": "Rincón de la Victoria",
    "torremolinos": "Torremolinos",
    "malaga": "Málaga",
}

# checkpoint / reanudar 
OUT_PATH_SUFFIX = "_filled.xlsx"
SAVE_ON_EACH_DISTRICT = True
SAVE_EVERY_DISTRICTS = 5
SAVE_EVERY_SECONDS = 180

FORCE_RECOMPUTE = False
SKIP_IF_ALL_QUARTERS_FILLED = True
# -----------


SPANISH_MONTHS = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12
}

MONTH_RE = re.compile(
    r"(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|setiembre|octubre|noviembre|diciembre)\s+(\d{4})",
    re.IGNORECASE
)

# pausas 
def human_pause(kind: str = "tiny"):
    """
    Pausas aleatorias para reducir ritmo de peticiones y parecer más humano.
    """
    ranges = {
        "tiny": (0.35, 0.90),   # entre selects
        "short": (0.8, 1.8),    # entre acciones dentro de la página
        "medium": (1.8, 3.5),   # tras navegación / cargas
        "long": (4.0, 8.0),     # entre distritos
    }
    a, b = ranges.get(kind, (0.8, 2.2))
    time.sleep(random.uniform(a, b))

def polite_sleep(a=0.8, b=2.2):
    time.sleep(random.uniform(a, b))

def backoff_sleep(attempt: int):
    base = min(60, (2 ** attempt))
    time.sleep(base + random.uniform(0.0, 1.5))


def parse_price_to_float(s: str) -> Optional[float]:
    if not s:
        return None
    s0 = s.strip().lower()
    if "n.d" in s0 or s0 == "nd":
        return None

    m = re.search(r"([\d\.]+(?:,\d+)?)\s*€\s*/\s*m2", s0)
    if not m:
        return None

    num = m.group(1).replace(".", "").replace(",", ".")
    try:
        return float(num)
    except ValueError:
        return None

def quarter_from_month(m: int) -> str:
    if m in (1, 2, 3): return "Q1"
    if m in (4, 5, 6): return "Q2"
    if m in (7, 8, 9): return "Q3"
    return "Q4"

def load_cod_mun_map(wb: openpyxl.Workbook) -> Dict[str, int]:
    ws = wb["COD_MUN"]
    out = {}
    for r in range(2, ws.max_row + 1):
        mun = ws.cell(r, 1).value
        code = ws.cell(r, 2).value
        if mun:
            out[str(mun)] = int(code) if code is not None else None
    return out

def build_row_index(ws_long) -> Dict[Tuple[str, str, int, str], int]:
    idx = {}
    for r in range(2, ws_long.max_row + 1):
        mun = ws_long.cell(r, 1).value
        dist = ws_long.cell(r, 2).value
        year = ws_long.cell(r, 3).value
        q = ws_long.cell(r, 4).value
        if mun and dist and year and q:
            idx[(str(mun), str(dist), int(year), str(q))] = r
    return idx

def get_needed_districts(ws_long, municipio: str) -> List[str]:
    dists = set()
    for r in range(2, ws_long.max_row + 1):
        mun = ws_long.cell(r, 1).value
        dist = ws_long.cell(r, 2).value
        if mun == municipio and dist:
            dists.add(str(dist))
    return sorted(dists)

def district_already_done(ws_long, row_index, municipio: str, distrito: str) -> bool:
    if FORCE_RECOMPUTE:
        return False

    any_filled = False
    all_filled = True

    for year in ANHOS:
        for q in TRIMESTRES:
            row = row_index.get((municipio, distrito, year, q))
            if not row:
                continue
            val = ws_long.cell(row, 5).value
            if val is not None:
                any_filled = True
            else:
                all_filled = False

    if SKIP_IF_ALL_QUARTERS_FILLED:
        return all_filled
    return any_filled


def click_best_effort(locator, page, timeout=20000) -> bool:
    try:
        locator.first.scroll_into_view_if_needed(timeout=timeout)
    except Exception:
        pass

    try:
        locator.first.click(timeout=timeout)
        return True
    except Exception:
        pass

    try:
        locator.first.click(timeout=timeout, force=True)
        return True
    except Exception:
        pass

    try:
        el = locator.first.element_handle()
        if el:
            page.evaluate(
                """(e) => {
                    e.scrollIntoView({block:'center'});
                    e.dispatchEvent(new MouseEvent('mouseover', {bubbles:true}));
                    e.dispatchEvent(new MouseEvent('mousemove', {bubbles:true}));
                    e.dispatchEvent(new MouseEvent('mousedown', {bubbles:true}));
                    e.dispatchEvent(new MouseEvent('mouseup', {bubbles:true}));
                    e.dispatchEvent(new MouseEvent('click', {bubbles:true}));
                }""",
                el
            )
            return True
    except Exception:
        pass

    return False

def click_consultar_informe(page, max_tries: int = 3):
    selectors = [
        "xpath=//button[contains(normalize-space(.),'Consultar informe')]",
        "xpath=//button[contains(normalize-space(.),'Consultar')]",
        "xpath=//a[contains(normalize-space(.),'Consultar informe')]",
        "xpath=//a[contains(normalize-space(.),'Consultar')]",
        "xpath=//input[((@type='submit' or @type='button') and contains(@value,'Consultar'))]",
    ]

    last_err = None
    for _ in range(max_tries):
        for sel in selectors:
            loc = page.locator(sel)
            if loc.count() > 0:
                human_pause("short")
                ok = click_best_effort(loc, page, timeout=30000)
                if ok:
                    try:
                        page.wait_for_load_state("domcontentloaded", timeout=30000)
                    except Exception:
                        pass
                    human_pause("medium")
                    return True
        last_err = "No encontré ningún elemento 'Consultar' (button/a/input)."
        human_pause("short")

    raise RuntimeError(last_err)

def click_ver_datos_mas_antiguos(page):
    try:
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    except Exception:
        try:
            page.mouse.wheel(0, 4000)
        except Exception:
            pass
    human_pause("short")

    link = page.locator(
        "xpath=//a[contains(.,'Ver datos más antiguos') or contains(.,'Ver datos mas antiguos')]"
    )
    try:
        link.first.wait_for(state="visible", timeout=15000)
    except Exception:
        pass

    link = page.locator(
        "xpath=//a[contains(.,'Ver datos más antiguos') or contains(.,'Ver datos mas antiguos')]"
    )
    if link.count() == 0:
        raise RuntimeError("No encuentro el enlace 'Ver datos más antiguos' (¿se cargó la tabla tras Consultar?)")

    human_pause("short")
    ok = click_best_effort(link, page, timeout=30000)
    if not ok:
        raise RuntimeError("No se pudo clicar 'Ver datos más antiguos'")

    try:
        page.wait_for_load_state("domcontentloaded", timeout=30000)
    except Exception:
        pass
    human_pause("medium")

def looks_blocked(page) -> bool:
    try:
        html = page.content().lower()
    except Exception:
        return False
    return any(k in html for k in [
        "access denied", "forbidden", "captcha", "are you a robot",
        "demasiadas solicitudes", "bloqueado", "verify you are human"
    ])


def select_option_slow(select_locator, *, label: str, kind_before="tiny", kind_after="tiny"):
    """
    Selecciona opción y mete pausas para bajar ritmo.
    """
    human_pause(kind_before)
    select_locator.select_option(label=label)
    human_pause(kind_after)


def extract_monthly_table_historico(page) -> List[Tuple[int, int, Optional[float]]]:
    table = page.locator("xpath=(//table)[1]")
    if table.count() == 0:
        return []

    rows = table.locator("xpath=.//tr[td]")
    dedup: Dict[Tuple[int, int], Optional[float]] = {}

    for i in range(rows.count()):
        tds = rows.nth(i).locator("td")
        if tds.count() < 2:
            continue

        mes_txt = tds.nth(0).inner_text().strip().lower()
        precio_txt = tds.nth(1).inner_text().strip().lower()

        m = MONTH_RE.search(mes_txt)
        if not m:
            continue

        month_name = m.group(1).lower()
        year = int(m.group(2))
        if year < 2015 or year > 2024:
            continue

        month = SPANISH_MONTHS.get(month_name)
        if not month:
            continue

        price = parse_price_to_float(precio_txt)
        dedup[(year, month)] = price

    return [(y, m, dedup[(y, m)]) for (y, m) in sorted(dedup.keys())]

def quarterly_from_monthly(monthly: List[Tuple[int, int, Optional[float]]]) -> Dict[Tuple[int, str], Optional[float]]:
    buckets: Dict[Tuple[int, str], List[float]] = {}

    for y, m, p in monthly:
        if p is None:
            continue
        buckets.setdefault((y, quarter_from_month(m)), []).append(p)

    out: Dict[Tuple[int, str], Optional[float]] = {}
    for y in ANHOS:
        for q in TRIMESTRES:
            vals = buckets.get((y, q), [])
            out[(y, q)] = (sum(vals) / len(vals)) if vals else None
    return out

def run(excel_path: str, headless: bool = True):
    out_path = excel_path.replace(".xlsx", OUT_PATH_SUFFIX)
    try:
        wb = openpyxl.load_workbook(out_path)
        print(f"[resume] Abriendo: {out_path}")
    except FileNotFoundError:
        wb = openpyxl.load_workbook(excel_path)
        print(f"[start] Abriendo: {excel_path}")

    ws_long = wb["LONG_INPUT"]
    cod_mun_map = load_cod_mun_map(wb)
    row_index = build_row_index(ws_long)

    processed_districts = 0
    last_save_time = time.time()

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=headless,
            args=["--disable-blink-features=AutomationControlled"],
        )
        context = browser.new_context(
            locale="es-ES",
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/122.0.0.0 Safari/537.36"),
            viewport={"width": 1366, "height": 768},
        )
        page = context.new_page()

        consecutive_errors = 0

        for municipio in MUNICIPIOS:
            dists = get_needed_districts(ws_long, municipio)
            if not dists:
                continue

            for distrito in dists:
                if district_already_done(ws_long, row_index, municipio, distrito):
                    print(f"==> {municipio} | {distrito}  [skip]")
                    continue

                print(f"==> {municipio} | {distrito}")

                try:
                    page.goto(URL, wait_until="domcontentloaded", timeout=60000)
                    human_pause("medium")

                    selects = page.locator("select")
                    selects.nth(0).wait_for(state="visible", timeout=20000)

                    # selects con pausas aleatorias
                    select_option_slow(selects.nth(0), label="Andalucía", kind_after="short")
                    select_option_slow(selects.nth(1), label="Málaga", kind_after="short")
                    select_option_slow(selects.nth(2), label=MUNICIPIO_LABEL.get(municipio, municipio), kind_after="short")

                    selects.nth(3).wait_for(state="visible", timeout=20000)
                    select_option_slow(selects.nth(3), label=distrito, kind_after="medium")

                    # Consultar -> Ver antiguos -> histórico
                    click_consultar_informe(page)
                    if looks_blocked(page):
                        raise RuntimeError("Posible bloqueo/captcha tras 'Consultar'.")

                    click_ver_datos_mas_antiguos(page)
                    if looks_blocked(page):
                        raise RuntimeError("Posible bloqueo/captcha al entrar en /historico/.")

                    monthly = extract_monthly_table_historico(page)
                    qmap = quarterly_from_monthly(monthly)

                    cod_mun = cod_mun_map.get(municipio)
                    updated = 0
                    for year in ANHOS:
                        for q in TRIMESTRES:
                            row = row_index.get((municipio, distrito, year, q))
                            if not row:
                                continue
                            if (not FORCE_RECOMPUTE) and ws_long.cell(row, 5).value is not None:
                                continue
                            price = qmap.get((year, q))
                            ws_long.cell(row, 5).value = (round(price, 2) if price is not None else None)
                            ws_long.cell(row, 6).value = cod_mun
                            updated += 1

                    print(f"    -> celdas actualizadas: {updated}")
                    consecutive_errors = 0

                except Exception as e:
                    consecutive_errors += 1
                    print(f"    [ERROR] {e}")

                    # screenshot debug
                    try:
                        safe_dist = re.sub(r"[^a-zA-Z0-9_-]+", "_", distrito)
                        page.screenshot(path=f"debug_{municipio}_{safe_dist}.png", full_page=True)
                        print(f"    [debug] screenshot: debug_{municipio}_{safe_dist}.png")
                    except Exception:
                        pass

                    backoff_sleep(min(consecutive_errors, 6))

                # checkpoint incremental
                processed_districts += 1
                need_save = False
                if SAVE_ON_EACH_DISTRICT:
                    need_save = True
                if processed_districts % SAVE_EVERY_DISTRICTS == 0:
                    need_save = True
                if time.time() - last_save_time >= SAVE_EVERY_SECONDS:
                    need_save = True

                if need_save:
                    wb.save(out_path)
                    last_save_time = time.time()
                    print(f"    [checkpoint] Guardado: {out_path}")

                # pausa entre distritos
                human_pause("long")

        context.close()
        browser.close()

    wb.save(out_path)
    print(f"\nOK. Guardado final: {out_path}")


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True, help="Ruta al Excel template (prices_template_districts.xlsx)")
    ap.add_argument("--headless", action="store_true", help="Ejecutar navegador en headless")
    args = ap.parse_args()
    run(args.excel, headless=args.headless)
